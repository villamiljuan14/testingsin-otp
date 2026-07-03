import logging

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Count, Q, Sum
from django.http import HttpResponse
from django.utils import timezone
from io import BytesIO
from ..models import Pedido, GuiaEnvio, EventoTracking, Reclamo, Hub, TipoServicio
from ..models.choices import EstadoPedido, EstadoGuia
from ..forms_tracking import EventoTrackingForm
from ..forms import GuiaEnvioForm, PedidoForm, HubForm, TipoServicioForm, ReclamoForm, ReclamoResolucionForm

logger = logging.getLogger(__name__)


def _filter_pedidos_queryset(request):
    """Lógica común de filtrado para lista y exportación."""
    user = request.user
    qs = Pedido.objects.select_related(
        'usuario', 'tipo_servicio', 'ciudad_origen', 'ciudad_destino', 'mensajero'
    )

    # Filtro por rol: admin vee todos, otros solo sus pedidos
    es_admin = user.is_staff or user.is_superuser
    if not es_admin and hasattr(user, 'rol') and user.rol:
        es_admin = str(user.rol.tipo_rol).upper() in ('ADMIN', 'ADMINISTRADOR')
    if not es_admin:
        qs = qs.filter(usuario=user)

    # Búsqueda full-text
    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(
            Q(numero_pedido__icontains=q) |
            Q(codigo_rastreo__icontains=q) |
            Q(nombre_destinatario__icontains=q) |
            Q(telefono_destinatario__icontains=q) |
            Q(descripcion_contenido__icontains=q) |
            Q(usuario__primer_nombre__icontains=q) |
            Q(usuario__primer_apellido__icontains=q)
        )

    # Filtro por estado
    estado = request.GET.get('estado', '').strip()
    if estado:
        qs = qs.filter(estado=estado)

    # Filtro por tipo de servicio
    servicio_id = request.GET.get('servicio', '').strip()
    if servicio_id and servicio_id.isdigit():
        qs = qs.filter(tipo_servicio_id=int(servicio_id))

    return qs


def _filter_hubs_queryset(request):
    """Lógica común de filtrado para lista de hubs e informes."""
    qs = Hub.objects.annotate(
        num_pedidos_origen=Count('pedidos_origen', distinct=True),
        num_pedidos_destino=Count('pedidos_destino', distinct=True)
    ).order_by('nombre')
    
    # Búsqueda avanzada
    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(
            Q(nombre__icontains=q) |
            Q(codigo__icontains=q) |
            Q(ciudad__icontains=q) |
            Q(departamento__icontains=q) |
            Q(direccion__icontains=q)
        )
        
    # Filtro por tipo de Hub
    tipo = request.GET.get('tipo', '').strip()
    if tipo:
        qs = qs.filter(tipo=tipo)
        
    # Filtro por estado operativo
    estado = request.GET.get('estado', '').strip()
    if estado == 'activo':
        qs = qs.filter(es_activo=True)
    elif estado == 'inactivo':
        qs = qs.filter(es_activo=False)
        
    return qs


def _filter_guias_queryset(request):
    """Lógica común de filtrado para lista de guías y exportación."""
    qs = GuiaEnvio.objects.select_related(
        'pedido', 'pedido__hub_origen', 'pedido__hub_destino', 'pedido__usuario'
    )
    
    # Búsqueda
    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(
            Q(numero_guia__icontains=q) |
            Q(pedido__numero_pedido__icontains=q) |
            Q(pedido__codigo_rastreo__icontains=q) |
            Q(pedido__nombre_destinatario__icontains=q) |
            Q(pedido__usuario__primer_nombre__icontains=q) |
            Q(pedido__usuario__primer_apellido__icontains=q)
        )
        
    # Filtro por estado
    estado = request.GET.get('estado', '').strip()
    if estado:
        qs = qs.filter(estado=estado)
        
    return qs.order_by('-fecha_generacion')


def _filter_hubs_queryset(request):
    """Lógica común de filtrado para hubs e informes."""
    qs = Hub.objects.annotate(
        num_pedidos_origen=Count('pedidos_origen', distinct=True),
        num_pedidos_destino=Count('pedidos_destino', distinct=True)
    ).order_by('nombre')
    
    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(
            Q(nombre__icontains=q) |
            Q(codigo__icontains=q) |
            Q(ciudad__icontains=q) |
            Q(departamento__icontains=q)
        )
    
    tipo = request.GET.get('tipo', '').strip()
    if tipo:
        qs = qs.filter(tipo=tipo)
        
    estado = request.GET.get('estado', '').strip()
    if estado == 'activo':
        qs = qs.filter(es_activo=True)
    elif estado == 'inactivo':
        qs = qs.filter(es_activo=False)
        
    return qs


def _filter_servicios_queryset(request):
    """Lógica común de filtrado para tipos de servicio."""
    qs = TipoServicio.objects.annotate(
        num_pedidos=Count('pedidos', distinct=True)
    ).order_by('nombre')
    
    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(Q(nombre__icontains=q) | Q(descripcion__icontains=q))
        
    return qs


def _filter_reclamos_queryset(request):
    """Lógica común de filtrado para reclamos e informes."""
    qs = Reclamo.objects.select_related(
        'pedido', 'reclamante', 'asignado_a'
    ).order_by('-fecha_radicacion')
    
    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(
            Q(numero_reclamo__icontains=q) |
            Q(pedido__numero_pedido__icontains=q) |
            Q(reclamante__primer_nombre__icontains=q) |
            Q(reclamante__email__icontains=q)
        )
        
    estado = request.GET.get('estado', '').strip()
    if estado:
        qs = qs.filter(estado=estado)
        
    prioridad = request.GET.get('prioridad', '').strip()
    if prioridad:
        qs = qs.filter(prioridad=prioridad)
        
    return qs


@login_required
def dashboard_pedidos_list(request):
    """Vista principal del CRUD de pedidos con filtros, paginación y stats."""
    pedidos = _filter_pedidos_queryset(request)

    # ── Stats ──
    total = Pedido.objects.count()
    en_transito = Pedido.objects.filter(
        estado__in=['EN_TRANSITO', 'EN_REPARTO', 'EN_HUB_ORIGEN', 'EN_HUB_DESTINO']
    ).count()
    entregados = Pedido.objects.filter(estado='ENTREGADO').count()
    cancelados = Pedido.objects.filter(estado='CANCELADO').count()

    stats = {
        'total': total,
        'en_transito': en_transito,
        'entregados': entregados,
        'cancelados': cancelados,
    }

    # ── Paginación ──
    items_por_pagina = int(request.GET.get('per_page', 10))
    if items_por_pagina not in [10, 20, 30, 40, 50]:
        items_por_pagina = 10

    paginator = Paginator(pedidos, items_por_pagina)
    page_number = request.GET.get('page', 1)
    try:
        page_obj = paginator.page(page_number)
    except (PageNotAnInteger, EmptyPage):
        page_obj = paginator.page(1)

    query_string = request.GET.copy()
    if 'page' in query_string:
        del query_string['page']
    query_params = query_string.urlencode()

    context = {
        'page_obj': page_obj,
        'stats': stats,
        'items_por_pagina': items_por_pagina,
        'query_params': query_params,
        'servicios': TipoServicio.objects.filter(es_activo=True),
        'estados_choices': [(e.value, e.label) for e in EstadoPedido],
        'current_q': request.GET.get('q', ''),
        'current_estado': request.GET.get('estado', ''),
        'current_servicio': request.GET.get('servicio', ''),
    }
    return render(request, 'dashboard/pedidos.html', context)


@login_required
def export_pedidos_view(request, file_format):
    """Exportar pedidos en PDF, XLSX o CSV aplicando los mismos filtros activos."""
    pedidos = _filter_pedidos_queryset(request)

    headers = ['N° Pedido', 'Código Rastreo', 'Cliente', 'Destinatario', 'Servicio',
               'Origen', 'Destino', 'Peso (kg)', 'Total', 'Estado', 'Fecha']

    if file_format == 'pdf':
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph,
            Spacer, HRFlowable, KeepTogether
        )
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm, cm
        from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT

        # ────────── PALETA DE COLORES PROFESIONAL ──────────
        C_BRAND       = colors.HexColor('#1B6BF5')   # Azul primario
        C_BRAND_DARK  = colors.HexColor('#1546C2')   # Azul oscuro
        C_DARK        = colors.HexColor('#0f172a')   # Slate-900
        C_LIGHT       = colors.HexColor('#f8fafc')   # Slate-50
        C_GRAY        = colors.HexColor('#e2e8f0')   # Slate-200
        C_GRAY_DARK   = colors.HexColor('#64748b')   # Slate-500
        C_TEXT        = colors.HexColor('#334155')   # Slate-700
        C_MUTED       = colors.HexColor('#94a3b8')   # Slate-400
        C_WHITE       = colors.white

        # Estados
        C_ENTREGADO   = colors.HexColor('#059669')   # Verde
        C_EN_TRANSITO = colors.HexColor('#d97706')   # Ámbar
        C_CANCELADO   = colors.HexColor('#dc2626')   # Rojo
        C_BORRADOR    = colors.HexColor('#6b7280')   # Gris

        # Fondos estado
        BG_ENTREGADO   = colors.HexColor('#d1fae5')  # Verde claro
        BG_EN_TRANSITO = colors.HexColor('#fef3c7')  # Ámbar claro
        BG_CANCELADO   = colors.HexColor('#fee2e2')  # Rojo claro
        BG_BORRADOR    = colors.HexColor('#f3f4f6')  # Gris claro

        PAGE_W, PAGE_H = landscape(A4)
        MARGIN = 1.3 * cm
        CONTENT_W = PAGE_W - (2 * MARGIN)

        buffer = BytesIO()

        # ────────── ESTILOS DE TEXTO ──────────
        styles = getSampleStyleSheet()
        
        s_title = ParagraphStyle(
            'title',
            fontName='Helvetica-Bold',
            fontSize=24,
            textColor=C_WHITE,
            leading=28,
            spaceAfter=0
        )
        
        s_subtitle = ParagraphStyle(
            'subtitle',
            fontName='Helvetica',
            fontSize=10,
            textColor=colors.HexColor('#bfdbfe'),
            leading=12,
            spaceAfter=0
        )
        
        s_header = ParagraphStyle(
            'header',
            fontName='Helvetica-Bold',
            fontSize=8.5,
            textColor=C_WHITE,
            leading=11,
            alignment=TA_CENTER,
            spaceAfter=0
        )
        
        s_cell = ParagraphStyle(
            'cell',
            fontName='Helvetica',
            fontSize=8,
            textColor=C_TEXT,
            leading=10,
            spaceAfter=0
        )
        
        s_cell_bold = ParagraphStyle(
            'cell_bold',
            fontName='Helvetica-Bold',
            fontSize=8,
            textColor=C_DARK,
            leading=10,
            spaceAfter=0
        )
        
        s_codigo = ParagraphStyle(
            'codigo',
            fontName='Courier',
            fontSize=7.5,
            textColor=C_BRAND,
            leading=9,
            spaceAfter=0
        )
        
        s_estado = ParagraphStyle(
            'estado',
            fontName='Helvetica-Bold',
            fontSize=7.5,
            textColor=C_DARK,
            leading=9,
            alignment=TA_CENTER,
            spaceAfter=0
        )

        # ────────── ESTADÍSTICAS ──────────
        total_qs = pedidos.count()
        entregado_qs = pedidos.filter(estado='ENTREGADO').count()
        en_transito_qs = pedidos.filter(
            estado__in=['EN_TRANSITO', 'EN_REPARTO', 'EN_HUB_ORIGEN', 'EN_HUB_DESTINO']
        ).count()
        cancelado_qs = pedidos.filter(estado='CANCELADO').count()
        total_valor = pedidos.aggregate(Sum('total_final'))['total_final__sum'] or 0

        # ────────── INFORMACIÓN DE GENERACIÓN ──────────
        gen_date = timezone.now().strftime('%d/%m/%Y %H:%M')
        filtros_str = ''
        if request.GET.get('q'):
            filtros_str += f'Búsqueda: "{request.GET["q"]}" | '
        if request.GET.get('estado'):
            filtros_str += f'Estado: {request.GET["estado"]} | '
        if request.GET.get('servicio'):
            filtros_str += f'Servicio ID: {request.GET["servicio"]} | '
        filtros_str = filtros_str.rstrip(' | ') or 'Sin filtros activos'

        # ────────── FUNCIÓN DE PÁGINA (Header/Footer) ──────────
        def on_page(canvas, doc):
            canvas.saveState()
            
            # ──── HEADER BAND (Logo + Título) ────
            header_h = 2.0 * cm
            canvas.setFillColor(C_BRAND)
            canvas.rect(0, PAGE_H - header_h, PAGE_W, header_h, fill=1, stroke=0)

            # Logo + Título
            canvas.setFillColor(C_WHITE)
            canvas.setFont('Helvetica-Bold', 20)
            canvas.drawString(MARGIN, PAGE_H - header_h + 1.05 * cm, 'ENVIART')
            
            canvas.setFillColor(colors.HexColor('#bfdbfe'))
            canvas.setFont('Helvetica', 9.5)
            canvas.drawString(MARGIN, PAGE_H - header_h + 0.45 * cm, 'Reporte de Gestión de Pedidos')

            # Metadata derecha
            canvas.setFont('Helvetica', 8)
            canvas.setFillColor(colors.HexColor('#bfdbfe'))
            canvas.drawRightString(
                PAGE_W - MARGIN, 
                PAGE_H - header_h + 1.05 * cm, 
                f'Generado: {gen_date}'
            )
            
            canvas.setFont('Helvetica', 7)
            canvas.drawRightString(
                PAGE_W - MARGIN, 
                PAGE_H - header_h + 0.45 * cm, 
                'Confidencial'
            )

            # ──── STATS BAND ────
            stats_top = PAGE_H - header_h
            stats_h = 1.8 * cm
            canvas.setFillColor(C_DARK)
            canvas.rect(0, stats_top - stats_h, PAGE_W, stats_h, fill=1, stroke=0)

            stat_items = [
                (str(total_qs), 'TOTAL', C_WHITE),
                (str(en_transito_qs), 'EN TRÁNSITO', C_ENTREGADO),
                (str(entregado_qs), 'ENTREGADOS', C_ENTREGADO),
                (str(cancelado_qs), 'CANCELADOS', C_CANCELADO),
                (f'${total_valor:,.0f}', 'VALOR TOTAL', C_BRAND),
            ]
            
            stat_width = PAGE_W / len(stat_items)
            
            for i, (val, lbl, col) in enumerate(stat_items):
                cx = i * stat_width + stat_width / 2
                
                # Número
                canvas.setFont('Helvetica-Bold', 18)
                canvas.setFillColor(col)
                canvas.drawCentredString(cx, stats_top - stats_h + 0.95 * cm, val)
                
                # Label
                canvas.setFont('Helvetica', 7)
                canvas.setFillColor(C_MUTED)
                canvas.drawCentredString(cx, stats_top - stats_h + 0.3 * cm, lbl)
                
                # Separador vertical
                if i > 0:
                    canvas.setStrokeColor(C_GRAY_DARK)
                    canvas.setLineWidth(0.3)
                    canvas.line(i * stat_width, stats_top - stats_h + 0.15 * cm,
                               i * stat_width, stats_top - stats_h + stats_h - 0.15 * cm)

            # ──── FILTROS INFO ────
            filters_top = stats_top - stats_h - 0.3 * cm
            canvas.setFont('Helvetica', 7)
            canvas.setFillColor(C_MUTED)
            canvas.drawString(MARGIN, filters_top, f'Filtros: {filtros_str}')

            # ──── FOOTER ────
            footer_h = 1.0 * cm
            canvas.setFillColor(C_LIGHT)
            canvas.rect(0, 0, PAGE_W, footer_h, fill=1, stroke=0)
            
            canvas.setStrokeColor(C_GRAY)
            canvas.setLineWidth(0.5)
            canvas.line(MARGIN, footer_h, PAGE_W - MARGIN, footer_h)
            
            canvas.setFont('Helvetica', 6.5)
            canvas.setFillColor(C_MUTED)
            canvas.drawString(
                MARGIN, 
                0.3 * cm, 
                'ENVIART Logistics · Sistema de Gestión de Pedidos · Confidencial'
            )
            canvas.drawRightString(
                PAGE_W - MARGIN, 
                0.3 * cm, 
                f'Página {doc.page}'
            )
            
            canvas.restoreState()

        # ────────── MAPEO ESTADO → COLORES ──────────
        ESTADO_COLORS = {
            'BORRADOR':       BG_BORRADOR,
            'CONFIRMADO':     colors.HexColor('#e0f2fe'),
            'EN_PREPARACION': colors.HexColor('#ede9fe'),
            'RECOLECTADO':    colors.HexColor('#f3e8ff'),
            'EN_HUB_ORIGEN':  colors.HexColor('#dbeafe'),
            'EN_HUB_DESTINO': colors.HexColor('#dbeafe'),
            'EN_TRANSITO':    BG_EN_TRANSITO,
            'EN_REPARTO':     BG_EN_TRANSITO,
            'ENTREGADO':      BG_ENTREGADO,
            'INTENTO_FALLIDO': colors.HexColor('#ffedd5'),
            'CANCELADO':      BG_CANCELADO,
            'DEVUELTO':       colors.HexColor('#fee2e2'),
        }

        ESTADO_TEXT_COLORS = {
            'BORRADOR': C_BORRADOR,
            'CONFIRMADO': colors.HexColor('#0369a1'),
            'EN_PREPARACION': colors.HexColor('#6d28d9'),
            'RECOLECTADO': colors.HexColor('#7c3aed'),
            'EN_HUB_ORIGEN': C_BRAND,
            'EN_HUB_DESTINO': C_BRAND,
            'EN_TRANSITO': C_EN_TRANSITO,
            'EN_REPARTO': C_EN_TRANSITO,
            'ENTREGADO': C_ENTREGADO,
            'INTENTO_FALLIDO': colors.HexColor('#ea580c'),
            'CANCELADO': C_CANCELADO,
            'DEVUELTO': colors.HexColor('#dc2626'),
        }

        # ────────── CONSTRUIR TABLA DE DATOS ──────────
        col_headers = [
            Paragraph('N° PEDIDO', s_header),
            Paragraph('CÓDIGO', s_header),
            Paragraph('CLIENTE', s_header),
            Paragraph('DESTINATARIO', s_header),
            Paragraph('SERVICIO', s_header),
            Paragraph('ORIGEN', s_header),
            Paragraph('DESTINO', s_header),
            Paragraph('PESO', s_header),
            Paragraph('VALOR', s_header),
            Paragraph('ESTADO', s_header),
            Paragraph('FECHA', s_header),
        ]

        # Anchos de columna proporcionales
        col_widths = [
            CONTENT_W * 0.10,  # N° Pedido
            CONTENT_W * 0.12,  # Código
            CONTENT_W * 0.11,  # Cliente
            CONTENT_W * 0.11,  # Destinatario
            CONTENT_W * 0.07,  # Servicio
            CONTENT_W * 0.10,  # Origen
            CONTENT_W * 0.10,  # Destino
            CONTENT_W * 0.06,  # Peso
            CONTENT_W * 0.07,  # Valor
            CONTENT_W * 0.09,  # Estado
            CONTENT_W * 0.07,  # Fecha
        ]

        table_data = [col_headers]
        row_estado_colors = []

        for p in pedidos:
            estado_label = p.get_estado_display()
            fragil_tag = ' ⚠' if p.es_fragil else ''
            peso_str = f'{float(p.peso_real_kg):.1f}'
            total_str = f'${float(p.total_final):,.0f}' if p.total_final else '—'
            
            origen = (p.ciudad_origen.nombre if p.ciudad_origen else '') or '—'
            destino = (p.ciudad_destino.nombre if p.ciudad_destino else '') or '—'
            cliente = (p.usuario.nombre_completo[:20] if p.usuario else '—')
            servicio = (p.tipo_servicio.nombre[:12] if p.tipo_servicio else '—')

            row = [
                Paragraph(p.numero_pedido or '—', s_codigo),
                Paragraph((p.codigo_rastreo or '—')[:14], s_codigo),
                Paragraph(cliente, s_cell),
                Paragraph((p.nombre_destinatario or '—')[:25] + fragil_tag, s_cell),
                Paragraph(servicio, s_cell),
                Paragraph(origen[:20], s_cell),
                Paragraph(destino[:20], s_cell),
                Paragraph(peso_str, s_cell_bold),
                Paragraph(total_str, s_cell_bold),
                Paragraph(estado_label, s_estado),
                Paragraph(p.fecha_pedido.strftime('%d/%m') if p.fecha_pedido else '—', s_cell),
            ]
            table_data.append(row)
            row_estado_colors.append(ESTADO_COLORS.get(p.estado, C_WHITE))

        # ────────── ESTILOS DE TABLA ──────────
        tbl_style = [
            # Encabezado
            ('BACKGROUND', (0, 0), (-1, 0), C_BRAND),
            ('TEXTCOLOR', (0, 0), (-1, 0), C_WHITE),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8.5),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            
            # Filas de datos
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            
            # Grid
            ('LINEBELOW', (0, 0), (-1, 0), 1.5, C_BRAND_DARK),
            ('LINEBELOW', (0, 1), (-1, -1), 0.3, C_GRAY),
            ('BOX', (0, 0), (-1, -1), 0.5, C_GRAY),
            
            # Alineación
            ('ALIGN', (7, 1), (8, -1), 'CENTER'),  # Peso y Valor
            ('ALIGN', (9, 1), (9, -1), 'CENTER'),  # Estado
            ('ALIGN', (10, 1), (10, -1), 'CENTER'),  # Fecha
        ]

        # Filas alternadas y colores por estado
        for row_idx in range(1, len(table_data)):
            # Fondo alterno
            if row_idx % 2 == 0:
                for col in range(len(col_headers)):
                    if col != 9:  # No colorear la columna de estado
                        tbl_style.append(('BACKGROUND', (col, row_idx), (col, row_idx), C_LIGHT))
            
            # Color de estado
            bg_estado = row_estado_colors[row_idx - 1]
            estado_text_color = ESTADO_TEXT_COLORS.get(table_data[row_idx][9].text, C_TEXT)
            tbl_style.append(('BACKGROUND', (9, row_idx), (9, row_idx), bg_estado))

        main_table = Table(table_data, colWidths=col_widths, repeatRows=1)
        main_table.setStyle(TableStyle(tbl_style))

        # ────────── CONSTRUIR DOCUMENTO ──────────
        top_offset = 2.0 * cm + 1.8 * cm + 0.5 * cm
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            leftMargin=MARGIN,
            rightMargin=MARGIN,
            topMargin=top_offset,
            bottomMargin=1.2 * cm,
        )
        
        story = [main_table]
        doc.build(story, onFirstPage=on_page, onLaterPages=on_page)

        buffer.seek(0)
        response = HttpResponse(buffer.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="enviart_pedidos_{timezone.now().strftime("%Y%m%d_%H%M")}.pdf"'
        return response

    elif file_format == 'xlsx':
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Pedidos'

        # Estilos
        header_fill = PatternFill(start_color='1B6BF5', end_color='1B6BF5', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=10)
        border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )

        # Headers
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Datos
        for row, p in enumerate(pedidos, 2):
            ws.cell(row=row, column=1).value = p.numero_pedido
            ws.cell(row=row, column=2).value = p.codigo_rastreo
            ws.cell(row=row, column=3).value = p.usuario.nombre_completo if p.usuario else '—'
            ws.cell(row=row, column=4).value = p.nombre_destinatario
            ws.cell(row=row, column=5).value = p.tipo_servicio.nombre if p.tipo_servicio else '—'
            ws.cell(row=row, column=6).value = (p.ciudad_origen.nombre if p.ciudad_origen else '') or '—'
            ws.cell(row=row, column=7).value = (p.ciudad_destino.nombre if p.ciudad_destino else '') or '—'
            ws.cell(row=row, column=8).value = float(p.peso_real_kg)
            ws.cell(row=row, column=9).value = float(p.total_final) if p.total_final else 0
            ws.cell(row=row, column=10).value = p.get_estado_display()
            ws.cell(row=row, column=11).value = p.fecha_pedido.strftime('%Y-%m-%d') if p.fecha_pedido else '—'

            # Aplicar estilos a filas
            for col in range(1, 12):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                if col in [8, 9]:
                    cell.alignment = Alignment(horizontal='right')
                else:
                    cell.alignment = Alignment(horizontal='left')
                if row % 2 == 0:
                    cell.fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')

        # Ajustar anchos de columna
        column_widths = [15, 18, 18, 18, 12, 15, 15, 10, 12, 15, 12]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="enviart_pedidos_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
        wb.save(response)
        return response

    elif file_format == 'csv':
        import csv
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="enviart_pedidos_{timezone.now().strftime("%Y%m%d_%H%M")}.csv"'
        response.write('\ufeff')  # BOM para Excel

        writer = csv.writer(response, delimiter=',', quoting=csv.QUOTE_ALL)
        writer.writerow(headers)
        
        for p in pedidos:
            writer.writerow([
                p.numero_pedido or '—',
                p.codigo_rastreo or '—',
                p.usuario.nombre_completo if p.usuario else '—',
                p.nombre_destinatario or '—',
                p.tipo_servicio.nombre if p.tipo_servicio else '—',
                (p.ciudad_origen.nombre if p.ciudad_origen else '') or '—',
                (p.ciudad_destino.nombre if p.ciudad_destino else '') or '—',
                f'{p.peso_real_kg:.1f}' if p.peso_real_kg else '—',
                f'{p.total_final:,.2f}' if p.total_final else '—',
                p.get_estado_display(),
                p.fecha_pedido.strftime('%Y-%m-%d') if p.fecha_pedido else '—',
            ])
        return response

    return redirect('dashboard_pedidos')


@login_required
def dashboard_guias_list(request):
    """Vista de guías de envío"""
    guias = _filter_guias_queryset(request).order_by('-fecha_generacion')
    
    # Estadísticas
    stats = {
        'total': GuiaEnvio.objects.count(),
        'hoy': GuiaEnvio.objects.filter(fecha_generacion__date=timezone.now().date()).count(),
        'en_transito': GuiaEnvio.objects.filter(estado='EN_TRANSITO').count(),
        'entregadas': GuiaEnvio.objects.filter(estado='ENTREGADA').count(),
    }
    
    # Paginación
    items_por_pagina = int(request.GET.get('per_page', 10))
    if items_por_pagina not in [10, 20, 30, 40, 50]:
        items_por_pagina = 10
    
    paginator = Paginator(guias, items_por_pagina)
    page_number = request.GET.get('page', 1)
    
    try:
        page_obj = paginator.page(page_number)
    except (PageNotAnInteger, EmptyPage):
        page_obj = paginator.page(1)
    
    context = {
        'page_obj': page_obj,
        'items_por_pagina': items_por_pagina,
        'stats': stats,
        'estados_guia': EstadoGuia.choices,
        'current_q': request.GET.get('q', ''),
        'current_estado': request.GET.get('estado', ''),
    }
    return render(request, 'dashboard/guias.html', context)


@login_required
def dashboard_tracking_list(request):
    """Vista de tracking y geolocalización con datatable"""
    tracking_events = EventoTracking.objects.select_related(
        'pedido', 'guia', 'hub', 'registrado_por'
    ).order_by('-fecha_registro')
    
    # Filtros
    query = request.GET.get('q', '').strip()
    tipo_evento = request.GET.get('tipo', '').strip()
    
    if query:
        tracking_events = tracking_events.filter(
            Q(pedido__numero_pedido__icontains=query) |
            Q(guia__numero_guia__icontains=query) |
            Q(ubicacion_texto__icontains=query) |
            Q(descripcion__icontains=query)
        )
    
    if tipo_evento:
        tracking_events = tracking_events.filter(tipo_evento=tipo_evento)
    
    # Estadísticas
    stats = {
        'total_eventos': EventoTracking.objects.count(),
        'hoy': EventoTracking.objects.filter(
            fecha_registro__date=timezone.now().date()
        ).count(),
    }
    
    # Tipos de evento para filtro
    tipos_evento = EventoTracking._meta.get_field('tipo_evento').choices
    
    # Paginación
    items_por_pagina = int(request.GET.get('per_page', 10))
    if items_por_pagina not in [10, 20, 30, 40, 50]:
        items_por_pagina = 10
    
    paginator = Paginator(tracking_events, items_por_pagina)
    page_number = request.GET.get('page', 1)
    try:
        page_obj = paginator.page(page_number)
    except (PageNotAnInteger, EmptyPage):
        page_obj = paginator.page(1)
    
    query_string = request.GET.copy()
    if 'page' in query_string:
        del query_string['page']
    query_params = query_string.urlencode()
    
    context = {
        'page_obj': page_obj,
        'items_por_pagina': items_por_pagina,
        'query_params': query_params,
        'stats': stats,
        'tipos_evento': tipos_evento,
        'current_tipo': tipo_evento,
        'current_q': query,
    }
    return render(request, 'dashboard/tracking.html', context)


@login_required
def dashboard_servicios_list(request):
    """Vista de tipos de servicio estandarizada"""
    servicios_qs = _filter_servicios_queryset(request)

    # Estadísticas básicas
    stats = {
        'total': TipoServicio.objects.count(),
        'uso_total': Pedido.objects.count(),
    }

    # Paginación
    items_por_pagina = int(request.GET.get('per_page', 10))
    if items_por_pagina not in [5, 10, 15, 20]:
        items_por_pagina = 10

    paginator = Paginator(servicios_qs, items_por_pagina)
    page_number = request.GET.get('page', 1)
    try:
        page_obj = paginator.page(page_number)
    except (PageNotAnInteger, EmptyPage):
        page_obj = paginator.page(1)

    query_string = request.GET.copy()
    query_string.pop('page', None)

    context = {
        'services': page_obj,
        'page_obj': page_obj,
        'stats': stats,
        'items_por_pagina': items_por_pagina,
        'query_params': query_string.urlencode(),
        'current_q': request.GET.get('q', ''),
    }
    return render(request, 'dashboard/servicios.html', context)


@login_required
def dashboard_hubs_list(request):
    """Vista de hubs y ubicaciones con estandarización Bento"""
    hubs_qs = _filter_hubs_queryset(request)
    
    # Estadísticas para el dashboard
    stats = {
        'total': Hub.objects.count(),
        'activos': Hub.objects.filter(es_activo=True).count(),
        'centrales': Hub.objects.filter(tipo='CENTRAL').count(),
        'regionales': Hub.objects.filter(tipo='REGIONAL').count(),
    }
    
    # Paginación
    items_por_pagina = int(request.GET.get('per_page', 10))
    if items_por_pagina not in [10, 20, 30, 40, 50]:
        items_por_pagina = 10
    
    paginator = Paginator(hubs_qs, items_por_pagina)
    page_number = request.GET.get('page', 1)
    
    try:
        page_obj = paginator.page(page_number)
    except (PageNotAnInteger, EmptyPage):
        page_obj = paginator.page(1)
        
    # Tipos de hub para el filtro
    tipos_hub = [
        ('CENTRAL', 'Hub Central'),
        ('REGIONAL', 'Hub Regional'),
        ('LOCAL', 'Hub Local'),
        ('PUNTO_ENTREGA', 'Punto de Entrega'),
    ]
    
    context = {
        'page_obj': page_obj,
        'items_por_pagina': items_por_pagina,
        'stats': stats,
        'tipos_hub': tipos_hub,
        'current_q': request.GET.get('q', ''),
        'current_tipo': request.GET.get('tipo', ''),
        'current_estado': request.GET.get('estado', ''),
    }
    return render(request, 'dashboard/hubs.html', context)


@login_required
def dashboard_reclamos_list(request):
    """Vista de reclamos estandarizada"""
    reclamos_qs = _filter_reclamos_queryset(request)
    
    # Estadísticas
    stats = {
        'total': Reclamo.objects.count(),
        'pendientes': Reclamo.objects.filter(estado='RADICADO').count(),
        'hoy': Reclamo.objects.filter(fecha_radicacion__date=timezone.now().date()).count(),
        'criticos': Reclamo.objects.filter(prioridad='CRITICO').count(),
    }
    
    # Paginación
    items_por_pagina = int(request.GET.get('per_page', 10))
    if items_por_pagina not in [10, 20, 30, 40, 50]:
        items_por_pagina = 10
    
    paginator = Paginator(reclamos_qs, items_por_pagina)
    page_number = request.GET.get('page', 1)
    
    try:
        page_obj = paginator.page(page_number)
    except (PageNotAnInteger, EmptyPage):
        page_obj = paginator.page(1)
        
    # Opciones para filtros
    estados_reclamo = [
        ('RADICADO', 'Radicado'),
        ('EN_REVISION', 'En revisión'),
        ('EN_INVESTIGACION', 'En investigación'),
        ('APROBADO', 'Aprobado'),
        ('RECHAZADO', 'Rechazado'),
    ]
    prioridades = [('BAJO', 'Bajo'), ('MEDIO', 'Medio'), ('ALTO', 'Alto'), ('CRITICO', 'Crítico')]
    
    context = {
        'page_obj': page_obj,
        'items_por_pagina': items_por_pagina,
        'stats': stats,
        'estados_reclamo': estados_reclamo,
        'prioridades': prioridades,
        'current_q': request.GET.get('q', ''),
        'current_estado': request.GET.get('estado', ''),
        'current_prioridad': request.GET.get('prioridad', ''),
    }
    return render(request, 'dashboard/reclamos.html', context)


# ═════════════════════════════════════════════════════════════════
# CRUD TRACKING VIEWS
# ═════════════════════════════════════════════════════════════════

@login_required
def tracking_create(request):
    """Crear nuevo evento de tracking"""
    if request.method == 'POST':
        logger.debug('tracking_create POST request; files=%s', request.FILES.keys())
        form = EventoTrackingForm(request.POST, request.FILES)
        if form.is_valid():
            evento = form.save(commit=False)
            evento.registrado_por = request.user

            # Guard adicional en la vista: verificar pedido entregado antes de guardar
            from ..models.choices import TipoEventoTracking as TET
            if EventoTracking.objects.filter(
                pedido=evento.pedido,
                tipo_evento=TET.ENTREGADO
            ).exists():
                messages.error(
                    request,
                    f'El pedido #{evento.pedido.numero_pedido} ya fue marcado como Entregado. '
                    f'No se pueden registrar más eventos de tracking.'
                )
                return render(request, 'dashboard/tracking_form.html', {
                    'form': form,
                    'title': 'Registrar Evento de Tracking',
                    'button_text': 'Registrar'
                })

            evento.save()
            messages.success(request, 'Evento de tracking registrado exitosamente.')
            return redirect('dashboard_tracking')
        else:
            logger.warning(
                'tracking_create form invalid; errors=%s; files=%s',
                form.errors.as_json(),
                list(request.FILES.keys())
            )
            messages.error(request, 'No se pudo guardar el evento. Revisa los errores del formulario.')
    else:
        form = EventoTrackingForm()
    
    return render(request, 'dashboard/tracking_form.html', {
        'form': form,
        'title': 'Registrar Evento de Tracking',
        'button_text': 'Registrar'
    })


@login_required
def tracking_update(request, pk):
    """Editar evento de tracking"""
    evento = get_object_or_404(EventoTracking, pk=pk)

    # 🔒 Eventos ENTREGADO son inmutables
    from ..models.choices import TipoEventoTracking as TET
    if evento.tipo_evento == TET.ENTREGADO:
        messages.error(
            request,
            f'El evento "{evento.get_tipo_evento_display()}" del pedido '
            f'#{evento.pedido.numero_pedido} corresponde a una entrega confirmada '
            f'y no puede ser modificado.'
        )
        return redirect('dashboard_tracking')

    if request.method == 'POST':
        logger.debug('tracking_update POST request; files=%s, evento_id=%s', request.FILES.keys(), pk)
        form = EventoTrackingForm(request.POST, request.FILES, instance=evento)
        if form.is_valid():
            form.save()
            messages.success(request, 'Evento de tracking actualizado exitosamente.')
            return redirect('dashboard_tracking')
        else:
            logger.warning(
                'tracking_update form invalid; errors=%s; files=%s; evento_id=%s',
                form.errors.as_json(),
                list(request.FILES.keys()),
                pk
            )
            messages.error(request, 'No se pudo actualizar el evento. Revisa los errores del formulario.')
    else:
        form = EventoTrackingForm(instance=evento)
    
    file_debug = None
    if evento and evento.evidencia_foto:
        file_debug = {
            'storage_class': evento.evidencia_foto.storage.__class__.__name__,
            'file_name': evento.evidencia_foto.name,
            'file_url': None,
        }
        try:
            file_debug['file_url'] = evento.evidencia_foto.url
        except Exception:
            file_debug['file_url'] = 'ERROR al generar URL'

    return render(request, 'dashboard/tracking_form.html', {
        'form': form,
        'title': 'Editar Evento de Tracking',
        'button_text': 'Actualizar',
        'evento': evento,
        'file_debug': file_debug,
    })


@login_required
def tracking_delete(request, pk):
    """Eliminar evento de tracking"""
    evento = get_object_or_404(EventoTracking, pk=pk)
    
    if request.method == 'POST':
        evento.delete()
        messages.success(request, 'Evento de tracking eliminado exitosamente.')
        return redirect('dashboard_tracking')
    
    return render(request, 'dashboard/tracking_confirm_delete.html', {
        'evento': evento
    })


@login_required
def export_guias_view(request, file_format):
    """Exportación multiformato para Guías de Envío"""
    guias = _filter_guias_queryset(request).order_by('-fecha_generacion')
    
    headers = [
        'No. Guía', 'No. Pedido', 'Remitente', 'Destinatario', 
        'Origen', 'Destino', 'Peso (kg)', 'Estado', 'Fecha Generación'
    ]

    if file_format == 'pdf':
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib.enums import TA_CENTER
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="enviart_guias_{timezone.now().strftime("%Y%m%d_%H%M")}.pdf"'
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        elements = []
        
        styles = getSampleStyleSheet()
        s_title = ParagraphStyle('title', fontName='Helvetica-Bold', fontSize=20, textColor=colors.HexColor('#1B6BF5'))
        s_header = ParagraphStyle('header', fontName='Helvetica-Bold', fontSize=9, textColor=colors.white, alignment=TA_CENTER)
        s_cell = ParagraphStyle('cell', fontName='Helvetica', fontSize=8.5)
        
        elements.append(Paragraph("Reporte de Guías de Envío - Enviart Logistics", s_title))
        elements.append(Spacer(1, 1*cm))
        
        data = [[Paragraph(h, s_header) for h in headers]]
        for g in guias:
            data.append([
                Paragraph(g.numero_guia, s_cell),
                Paragraph(g.pedido.numero_pedido, s_cell),
                Paragraph(f"{g.pedido.usuario.primer_nombre} {g.pedido.usuario.primer_apellido}" if g.pedido.usuario else "—", s_cell),
                Paragraph(g.pedido.nombre_destinatario, s_cell),
                Paragraph(g.pedido.hub_origen.nombre if g.pedido.hub_origen else "—", s_cell),
                Paragraph(g.pedido.hub_destino.nombre if g.pedido.hub_destino else "—", s_cell),
                Paragraph(f"{g.peso_final_kg:.2f}", s_cell),
                Paragraph(f"{g.get_estado_display()}", s_cell),
                Paragraph(g.fecha_generacion.strftime('%d/%m/%Y'), s_cell),
            ])
            
        t = Table(data, repeatRows=1, colWidths=[3.5*cm, 3*cm, 4*cm, 4*cm, 3.5*cm, 3.5*cm, 2*cm, 2.5*cm, 2.5*cm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1B6BF5')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(t)
        doc.build(elements)
        response.write(buffer.getvalue())
        buffer.close()
        return response

    elif file_format == 'xlsx':
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Guias'
        
        header_fill = PatternFill(start_color='1B6BF5', end_color='1B6BF5', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            
        for row, g in enumerate(guias, 2):
            ws.cell(row=row, column=1).value = g.numero_guia
            ws.cell(row=row, column=2).value = g.pedido.numero_pedido
            ws.cell(row=row, column=3).value = f"{g.pedido.usuario.primer_nombre} {g.pedido.usuario.primer_apellido}" if g.pedido.usuario else "—"
            ws.cell(row=row, column=4).value = g.pedido.nombre_destinatario
            ws.cell(row=row, column=5).value = g.pedido.hub_origen.nombre if g.pedido.hub_origen else "—"
            ws.cell(row=row, column=6).value = g.pedido.hub_destino.nombre if g.pedido.hub_destino else "—"
            ws.cell(row=row, column=7).value = float(g.peso_final_kg)
            ws.cell(row=row, column=8).value = g.get_estado_display()
            ws.cell(row=row, column=9).value = g.fecha_generacion.strftime('%Y-%m-%d')
            
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="enviart_guias_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
        wb.save(response)
        return response

    elif file_format == 'csv':
        import csv
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="enviart_guias_{timezone.now().strftime("%Y%m%d_%H%M")}.csv"'
        response.write('\ufeff')
        writer = csv.writer(response)
        writer.writerow(headers)
        for g in guias:
            writer.writerow([
                g.numero_guia, g.pedido.numero_pedido,
                f"{g.pedido.usuario.primer_nombre} {g.pedido.usuario.primer_apellido}" if g.pedido.usuario else "—",
                g.pedido.nombre_destinatario,
                g.pedido.hub_origen.nombre if g.pedido.hub_origen else "—",
                g.pedido.hub_destino.nombre if g.pedido.hub_destino else "—",
                g.peso_final_kg,
                g.get_estado_display(),
                g.fecha_generacion.strftime('%Y-%m-%d'),
            ])
        return response

    return redirect('dashboard_guias')


@login_required
def export_hubs_view(request, file_format):
    """Exportación multiformato para Hubs"""
    hubs = _filter_hubs_queryset(request)
    headers = ['Código', 'Nombre', 'Tipo', 'Ciudad', 'Departamento', 'Dirección', 'Teléfono', 'Activo']

    if file_format == 'pdf':
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib.enums import TA_CENTER
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="enviart_hubs_{timezone.now().strftime("%Y%m%d_%H%M")}.pdf"'
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        elements = []
        
        styles = getSampleStyleSheet()
        s_title = ParagraphStyle('title', fontName='Helvetica-Bold', fontSize=20, textColor=colors.HexColor('#1B6BF5'))
        s_header = ParagraphStyle('header', fontName='Helvetica-Bold', fontSize=9, textColor=colors.white, alignment=TA_CENTER)
        s_cell = ParagraphStyle('cell', fontName='Helvetica', fontSize=8)
        
        elements.append(Paragraph("Reporte de Hubs Logísticos - Enviart", s_title))
        elements.append(Spacer(1, 1*cm))
        
        data = [[Paragraph(h, s_header) for h in headers]]
        for h in hubs:
            data.append([
                Paragraph(h.codigo, s_cell),
                Paragraph(h.nombre, s_cell),
                Paragraph(h.get_tipo_display(), s_cell),
                Paragraph(h.ciudad, s_cell),
                Paragraph(h.departamento, s_cell),
                Paragraph(h.direccion, s_cell),
                Paragraph(h.telefono, s_cell),
                Paragraph("Sí" if h.es_activo else "No", s_cell),
            ])
            
        t = Table(data, repeatRows=1, colWidths=[2.5*cm, 3.5*cm, 3*cm, 2.5*cm, 3*cm, 5*cm, 3*cm, 2*cm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1B6BF5')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(t)
        doc.build(elements)
        response.write(buffer.getvalue())
        buffer.close()
        return response

    elif file_format == 'xlsx':
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Hubs'
        for col, head in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=head)
        for row, h in enumerate(hubs, 2):
            ws.cell(row=row, column=1).value = h.codigo
            ws.cell(row=row, column=2).value = h.nombre
            ws.cell(row=row, column=3).value = h.get_tipo_display()
            ws.cell(row=row, column=4).value = h.ciudad
            ws.cell(row=row, column=5).value = h.departamento
            ws.cell(row=row, column=6).value = h.direccion
            ws.cell(row=row, column=7).value = h.telefono
            ws.cell(row=row, column=8).value = "Activo" if h.es_activo else "Inactivo"
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="enviart_hubs_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
        wb.save(response)
        return response

    import csv
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="enviart_hubs_{timezone.now().strftime("%Y%m%d_%H%M")}.csv"'
    writer = csv.writer(response)
    writer.writerow(headers)
    for h in hubs:
        writer.writerow([h.codigo, h.nombre, h.get_tipo_display(), h.ciudad, h.departamento, h.direccion, h.telefono, h.es_activo])
    return response


@login_required
def export_servicios_view(request, file_format):
    """Exportación multiformato para Servicios"""
    servicios = _filter_servicios_queryset(request)
    headers = ['Nombre', 'Descripción', 'Precio Base', 'Pedidos Asignados']

    if file_format == 'xlsx':
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Servicios'
        for col, head in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=head)
        for row, s in enumerate(servicios, 2):
            ws.cell(row=row, column=1).value = s.nombre
            ws.cell(row=row, column=2).value = s.descripcion
            ws.cell(row=row, column=3).value = float(s.precio_base)
            ws.cell(row=row, column=4).value = s.num_pedidos
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="enviart_servicios_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
        wb.save(response)
        return response

    import csv
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="enviart_servicios_{timezone.now().strftime("%Y%m%d_%H%M")}.csv"'
    writer = csv.writer(response)
    writer.writerow(headers)
    for s in servicios:
        writer.writerow([s.nombre, s.descripcion, s.precio_base, s.num_pedidos])
    return response


@login_required
def export_reclamos_view(request, file_format):
    """Exportación multiformato para Reclamos"""
    reclamos = _filter_reclamos_queryset(request)
    headers = ['No. Reclamo', 'No. Pedido', 'Reclamante', 'Tipo', 'Estado', 'Prioridad', 'Fecha Radicación', 'Valor Reclamado']

    if file_format == 'pdf':
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="enviart_reclamos_{timezone.now().strftime("%Y%m%d_%H%M")}.pdf"'
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        elements = []
        
        styles = getSampleStyleSheet()
        s_title = ParagraphStyle('title', fontName='Helvetica-Bold', fontSize=20, textColor=colors.HexColor('#E11D48'))
        s_header = ParagraphStyle('header', fontName='Helvetica-Bold', fontSize=9, textColor=colors.white)
        s_cell = ParagraphStyle('cell', fontName='Helvetica', fontSize=8)
        
        elements.append(Paragraph("Reporte de Reclamos e Incidencias - Enviart", s_title))
        elements.append(Spacer(1, 1*cm))
        
        data = [[Paragraph(h, s_header) for h in headers]]
        for r in reclamos:
            data.append([
                Paragraph(r.numero_reclamo, s_cell),
                Paragraph(r.pedido.numero_pedido, s_cell),
                Paragraph(r.reclamante.email, s_cell),
                Paragraph(r.get_tipo_display(), s_cell),
                Paragraph(r.get_estado_display(), s_cell),
                Paragraph(r.get_prioridad_display(), s_cell),
                Paragraph(r.fecha_radicacion.strftime('%d/%m/%Y'), s_cell),
                Paragraph(f"${r.valor_reclamado:,.2f}", s_cell),
            ])
            
        t = Table(data, repeatRows=1, colWidths=[3.5*cm, 3*cm, 6*cm, 4*cm, 3.5*cm, 2.5*cm, 3*cm, 3.5*cm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E11D48')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(t)
        doc.build(elements)
        response.write(buffer.getvalue())
        buffer.close()
        return response

    elif file_format == 'xlsx':
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Reclamos'
        for col, head in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=head)
        for row, r in enumerate(reclamos, 2):
            ws.cell(row=row, column=1).value = r.numero_reclamo
            ws.cell(row=row, column=2).value = r.pedido.numero_pedido
            ws.cell(row=row, column=3).value = r.reclamante.email
            ws.cell(row=row, column=4).value = r.get_tipo_display()
            ws.cell(row=row, column=5).value = r.get_estado_display()
            ws.cell(row=row, column=6).value = r.get_prioridad_display()
            ws.cell(row=row, column=7).value = r.fecha_radicacion.strftime('%Y-%m-%d')
            ws.cell(row=row, column=8).value = float(r.valor_reclamado)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="enviart_reclamos_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
        wb.save(response)
        return response

    import csv
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="enviart_reclamos_{timezone.now().strftime("%Y%m%d_%H%M")}.csv"'
    writer = csv.writer(response)
    writer.writerow(headers)
    for r in reclamos:
        writer.writerow([r.numero_reclamo, r.pedido.numero_pedido, r.reclamante.email, r.get_tipo_display(), r.get_estado_display(), r.get_prioridad_display(), r.fecha_radicacion, r.valor_reclamado])
    return response


@login_required
def export_tracking_view(request, file_format):
    """Exportación multiformato para Eventos de Tracking"""
    eventos = EventoTracking.objects.select_related(
        'pedido', 'guia', 'registrado_por'
    ).order_by('-fecha_registro')

    q = request.GET.get('q', '').strip()
    tipo = request.GET.get('tipo', '').strip()
    if q:
        from django.db.models import Q
        eventos = eventos.filter(
            Q(pedido__numero_pedido__icontains=q) |
            Q(guia__numero_guia__icontains=q) |
            Q(ubicacion_texto__icontains=q)
        )
    if tipo:
        eventos = eventos.filter(tipo_evento=tipo)

    headers = ['ID', 'Pedido', 'Guía', 'Tipo Evento', 'Ubicación', 'Latitud', 'Longitud', 'Registrado Por', 'Fecha']

    if file_format == 'pdf':
        rows = []
        for e in eventos:
            rows.append([
                e.id,
                e.pedido.numero_pedido if e.pedido else '—',
                e.guia.numero_guia if e.guia else '—',
                e.get_tipo_evento_display(),
                e.ubicacion_texto or '—',
                str(e.latitud) if e.latitud else '—',
                str(e.longitud) if e.longitud else '—',
                e.registrado_por.email if e.registrado_por else 'Sistema',
                e.fecha_registro.strftime('%Y-%m-%d %H:%M') if e.fecha_registro else '—',
            ])
        from apps.core.views import render_to_pdf
        return render_to_pdf(headers, rows, "REPORTE DE TRACKING", f"enviart_tracking_{timezone.now().strftime('%Y%m%d_%H%M')}.pdf")

    import tablib
    dataset = tablib.Dataset(headers=headers)
    for e in eventos:
        dataset.append([
            e.id,
            e.pedido.numero_pedido if e.pedido else '—',
            e.guia.numero_guia if e.guia else '—',
            e.get_tipo_evento_display(),
            e.ubicacion_texto or '—',
            float(e.latitud) if e.latitud else None,
            float(e.longitud) if e.longitud else None,
            e.registrado_por.email if e.registrado_por else 'Sistema',
            e.fecha_registro.strftime('%Y-%m-%d %H:%M') if e.fecha_registro else '—',
        ])
    from apps.core.views import export_dataset
    return export_dataset(dataset, file_format, f'enviart_tracking_{timezone.now().strftime("%Y%m%d_%H%M")}')


# ═════════════════════════════════════════════════════════════════
# CRUD GUÍAS DE ENVÍO
# ═════════════════════════════════════════════════════════════════

@login_required
def guia_create_dashboard_view(request, pedido_id):
    """Generar una guía para un pedido desde el dashboard."""
    pedido = get_object_or_404(Pedido, id=pedido_id)
    
    # Verificar si ya existe una guía
    if hasattr(pedido, 'guia'):
        messages.warning(request, f'El pedido #{pedido.numero_pedido} ya tiene una guía asignada.')
        return redirect('dashboard_guias')
    
    if request.method == 'POST':
        form = GuiaEnvioForm(request.POST)
        if form.is_valid():
            guia = form.save(commit=False)
            guia.pedido = pedido
            guia.generada_por = request.user
            
            # Autogenerar dimensiones si no vienen
            if not guia.dimensiones:
                guia.dimensiones = f"{pedido.largo_cm}x{pedido.ancho_cm}x{pedido.alto_cm} cm"
            
            guia.save()
            
            # Registrar evento de tracking
            from ..models import EventoTracking, TipoEventoTracking
            EventoTracking.objects.create(
                pedido=pedido,
                tipo_evento=TipoEventoTracking.GUIA_CREADA,
                ubicacion_texto=pedido.hub_origen.nombre if pedido.hub_origen else "Centro de Acopio",
                registrado_por=request.user,
                descripcion='Guía de envío generada administrativamente.'
            )
            
            messages.success(request, f'Guía {guia.numero_guia} generada exitosamente.')
            return redirect('dashboard_guias')
    else:
        # Pre-poblar datos sugeridos
        initial_data = {
            'peso_final_kg': pedido.peso_real_kg,
            'dimensiones': f"{pedido.largo_cm}x{pedido.ancho_cm}x{pedido.alto_cm} cm" if pedido.largo_cm else "",
            'es_fragil': pedido.es_fragil,
            'requiere_firma': pedido.requiere_firma
        }
        form = GuiaEnvioForm(initial=initial_data)
    
    return render(request, 'dashboard/guia_form.html', {
        'form': form,
        'pedido': pedido,
        'title': 'Generar Guía de Envío',
        'button_text': 'Generar Documento'
    })


@login_required
def guia_update_view(request, pk):
    """Editar una guía existente."""
    guia = get_object_or_404(GuiaEnvio, pk=pk)
    
    if request.method == 'POST':
        form = GuiaEnvioForm(request.POST, instance=guia)
        if form.is_valid():
            form.save()
            messages.success(request, f'Guía {guia.numero_guia} actualizada correctamente.')
            return redirect('dashboard_guias')
    else:
        form = GuiaEnvioForm(instance=guia)
        
    return render(request, 'dashboard/guia_form.html', {
        'form': form,
        'guia': guia,
        'pedido': guia.pedido,
        'title': 'Editar Guía de Envío',
        'button_text': 'Guardar Cambios'
    })


@login_required
def guia_delete_view(request, pk):
    """Eliminar (cancelar) una guía."""
    guia = get_object_or_404(GuiaEnvio, pk=pk)
    numero = guia.numero_guia
    
    if request.method == 'POST':
        guia.delete()
        messages.success(request, f'La guía {numero} ha sido eliminada. El pedido ahora puede recibir una nueva guía.')
        return redirect('dashboard_guias')
        
    return render(request, 'dashboard/tracking_confirm_delete.html', {
        'objeto': guia,
        'nombre_objeto': f'Guía {numero}',
        'url_cancel': 'dashboard_guias'
    })


# ═════════════════════════════════════════════════════════════════
# CRUD HUBS LOGÍSTICOS
# ═════════════════════════════════════════════════════════════════

@login_required
def hub_create_view(request):
    """Crear un nuevo Hub logístico."""
    if request.method == 'POST':
        form = HubForm(request.POST)
        if form.is_valid():
            hub = form.save()
            messages.success(request, f'Hub {hub.codigo} creado exitosamente.')
            return redirect('dashboard_hubs')
    else:
        form = HubForm()
    
    return render(request, 'dashboard/hub_form.html', {
        'form': form,
        'title': 'Registrar Nuevo Hub',
        'button_text': 'Crear Hub'
    })


@login_required
def hub_update_view(request, pk):
    """Editar un Hub existente."""
    hub = get_object_or_404(Hub, pk=pk)
    
    if request.method == 'POST':
        form = HubForm(request.POST, instance=hub)
        if form.is_valid():
            form.save()
            messages.success(request, f'Hub {hub.codigo} actualizado correctamente.')
            return redirect('dashboard_hubs')
    else:
        form = HubForm(instance=hub)
        
    return render(request, 'dashboard/hub_form.html', {
        'form': form,
        'hub': hub,
        'title': 'Editar Hub Logístico',
        'button_text': 'Guardar Cambios'
    })


@login_required
def hub_delete_view(request, pk):
    """Eliminar (desactivar) un Hub."""
    hub = get_object_or_404(Hub, pk=pk)
    codigo = hub.codigo
    
    if request.method == 'POST':
        hub.delete()
        messages.success(request, f'El Hub {codigo} ha sido eliminado del sistema.')
        return redirect('dashboard_hubs')
        
    return render(request, 'dashboard/tracking_confirm_delete.html', {
        'objeto': hub,
        'nombre_objeto': f'Hub {codigo} ({hub.nombre})',
        'url_cancel': 'dashboard_hubs'
    })


# ═════════════════════════════════════════════════════════════════
# CRUD TIPOS DE SERVICIO
# ═════════════════════════════════════════════════════════════════

@login_required
def servicio_create_view(request):
    """Crear un nuevo tipo de servicio."""
    if request.method == 'POST':
        form = TipoServicioForm(request.POST)
        if form.is_valid():
            servicio = form.save()
            messages.success(request, f'Servicio {servicio.nombre} creado exitosamente.')
            return redirect('dashboard_servicios')
    else:
        form = TipoServicioForm()
    
    return render(request, 'dashboard/servicio_form.html', {
        'form': form,
        'title': 'Configurar Nuevo Servicio',
        'button_text': 'Crear Servicio'
    })


@login_required
def servicio_update_view(request, pk):
    """Editar un servicio existente."""
    servicio = get_object_or_404(TipoServicio, pk=pk)
    
    if request.method == 'POST':
        form = TipoServicioForm(request.POST, instance=servicio)
        if form.is_valid():
            form.save()
            messages.success(request, f'Servicio {servicio.nombre} actualizado correctamente.')
            return redirect('dashboard_servicios')
    else:
        form = TipoServicioForm(instance=servicio)
        
    return render(request, 'dashboard/servicio_form.html', {
        'form': form,
        'servicio': servicio,
        'title': 'Editar Tipo de Servicio',
        'button_text': 'Guardar Cambios'
    })


@login_required
def servicio_delete_view(request, pk):
    """Eliminar (desactivar) un servicio."""
    servicio = get_object_or_404(TipoServicio, pk=pk)
    nombre = servicio.nombre
    
    if request.method == 'POST':
        servicio.delete()
        messages.success(request, f'El servicio {nombre} ha sido eliminado.')
        return redirect('dashboard_servicios')
        
    return render(request, 'dashboard/tracking_confirm_delete.html', {
        'objeto': servicio,
        'nombre_objeto': f'Servicio {nombre}',
        'url_cancel': 'dashboard_servicios'
    })


# ═════════════════════════════════════════════════════════════════
# CRUD RECLAMOS
# ═════════════════════════════════════════════════════════════════

@login_required
def reclamo_create_view(request):
    """Radicar un nuevo reclamo administrativamente."""
    if request.method == 'POST':
        form = ReclamoForm(request.POST)
        if form.is_valid():
            reclamo = form.save(commit=False)
            reclamo.reclamante = request.user  # Administrativamente se marca quien radica
            reclamo.save()
            messages.success(request, f'Reclamo {reclamo.numero_reclamo} radicado exitosamente.')
            return redirect('dashboard_reclamos')
    else:
        # Si viene un pedido por GET, pre-seleccionarlo
        pedido_id = request.GET.get('pedido')
        initial = {}
        if pedido_id:
            pedido = get_object_or_404(Pedido, id=pedido_id)
            initial = {
                'pedido': pedido,
                'nombre_reclamante': f"{pedido.usuario.primer_nombre} {pedido.usuario.primer_apellido}",
                'email_reclamante': pedido.usuario.email,
                'telefono_reclamante': pedido.usuario.telefono or ""
            }
        form = ReclamoForm(initial=initial)
    
    return render(request, 'dashboard/reclamo_form.html', {
        'form': form,
        'title': 'Radicar Nuevo Reclamo',
        'button_text': 'Radicar Reclamo'
    })


@login_required
def reclamo_update_view(request, pk):
    """Gestión administrativa (Resolución) de un reclamo."""
    reclamo = get_object_or_404(Reclamo, pk=pk)
    
    if request.method == 'POST':
        form = ReclamoResolucionForm(request.POST, instance=reclamo)
        if form.is_valid():
            form.save()
            messages.success(request, f'Resolución del reclamo {reclamo.numero_reclamo} guardada.')
            return redirect('dashboard_reclamos')
    else:
        form = ReclamoResolucionForm(instance=reclamo)
        
    return render(request, 'dashboard/reclamo_form.html', {
        'form': form,
        'reclamo': reclamo,
        'title': 'Gestionar Reclamo',
        'button_text': 'Actualizar Resolución'
    })


@login_required
def reclamo_delete_view(request, pk):
    """Eliminar un reclamo."""
    reclamo = get_object_or_404(Reclamo, pk=pk)
    numero = reclamo.numero_reclamo
    
    if request.method == 'POST':
        reclamo.delete()
        messages.success(request, f'El reclamo {numero} ha sido eliminado.')
        return redirect('dashboard_reclamos')
        
    return render(request, 'dashboard/tracking_confirm_delete.html', {
        'objeto': reclamo,
        'nombre_objeto': f'Reclamo {numero}',
        'url_cancel': 'dashboard_reclamos'
    })
