from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.views.decorators.http import require_POST
from django.views.generic import View
from django.contrib.auth.decorators import login_required, permission_required
from django.db.models import Q
from apps.core.decorators import admin_or_permission
from django.http import HttpResponse
from django.utils import timezone
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from ..models import Usuario, Rol, TipoRol, TipoDocumento
from ..forms import UsuarioForm
import tablib
from io import BytesIO

# ── Helpers de Filtrado ──

def _filter_usuarios_queryset(request):
    """Lógica común de filtrado para lista y exportación."""
    usuarios = Usuario.objects.select_related('rol').all()
    
    # Búsqueda
    q = request.GET.get('q', '').strip()
    if q:
        usuarios = usuarios.filter(
            Q(primer_nombre__icontains=q) |
            Q(segundo_nombre__icontains=q) |
            Q(primer_apellido__icontains=q) |
            Q(segundo_apellido__icontains=q) |
            Q(email__icontains=q) |
            Q(doc_usuario__icontains=q) |
            Q(telefono__icontains=q)
        )
    
    # Filtro por rol
    rol_id = request.GET.get('rol', '').strip()
    if rol_id and rol_id.isdigit():
        usuarios = usuarios.filter(rol_id=int(rol_id))
    
    # Filtro por estado
    estado = request.GET.get('estado', '').strip()
    if estado == 'activo':
        usuarios = usuarios.filter(is_active=True)
    elif estado == 'inactivo':
        usuarios = usuarios.filter(is_active=False)
    
    return usuarios.order_by('-created_at')


# ── Vistas de Autenticación Pública ──

def login_view(request):
    """Muestra el formulario de login y procesa las credenciales."""
    if request.user.is_authenticated:
        # Redirigir según el rol del usuario ya autenticado
        if request.user.rol:
            role_val = str(request.user.rol.tipo_rol).upper()
            if role_val in (TipoRol.ADMINISTRADOR, TipoRol.ADMINISTRADOR.label, 'ADMINISTRADOR'):
                return redirect('dashboard_index')
            elif role_val in (TipoRol.CLIENTE, TipoRol.CLIENTE.label):
                return redirect('dashboard_cliente')
            elif role_val in (TipoRol.MENSAJERO, TipoRol.MENSAJERO.label):
                return redirect('dashboard_mensajero')
            elif role_val in (TipoRol.PROVEEDOR, TipoRol.PROVEEDOR.label):
                return redirect('dashboard_proveedor')
        return redirect('dashboard_index')

    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '')

        if not email or not password:
            messages.error(request, 'Email y contraseña son obligatorios.')
            return render(request, 'login.html')

        email = Usuario.objects.normalize_email(email)
        try:
            db_user = Usuario.objects.get(email__iexact=email)
            email = db_user.email
        except Usuario.DoesNotExist:
            pass

        user = authenticate(request, username=email, password=password)

        if user is not None:
            if not user.is_active:
                messages.error(request, 'Tu cuenta está inactiva. Contacta al administrador.')
                return render(request, 'login.html')

            if not request.POST.get('remember_me'):
                request.session.set_expiry(0)

            login(request, user)
            
            # Redirección basada en Rol
            if user.rol:
                role_val = str(user.rol.tipo_rol).upper()
                if role_val in (TipoRol.ADMINISTRADOR, TipoRol.ADMINISTRADOR.label, 'ADMINISTRADOR'):
                    return redirect('dashboard_index')  # KPI dashboard + pedidos
                elif role_val in (TipoRol.CLIENTE, TipoRol.CLIENTE.label):
                    return redirect('dashboard_cliente')
                elif role_val in (TipoRol.MENSAJERO, TipoRol.MENSAJERO.label):
                    return redirect('dashboard_mensajero')
                elif role_val in (TipoRol.PROVEEDOR, TipoRol.PROVEEDOR.label):
                    return redirect('dashboard_proveedor')

            # Redirección por defecto si no encaja: KPI dashboard
            return redirect('dashboard_index')
        else:
            return render(request, 'login.html', {'error_manual': True})

    return render(request, 'login.html')


def register_view(request):
    """Procesa el formulario de registro de 3 pasos."""
    if request.user.is_authenticated:
        return redirect('home')

    if request.method == 'POST':
        tipo_documento = request.POST.get('tipoDocumento', '').strip()
        doc_usuario = request.POST.get('docUsuario', '').strip()
        primer_nombre = request.POST.get('primerNombre', '').strip()
        segundo_nombre = request.POST.get('segundoNombre', '').strip() or None
        primer_apellido = request.POST.get('primerApellido', '').strip()
        segundo_apellido = request.POST.get('segundoApellido', '').strip() or None
        codigo_pais = request.POST.get('codigoPais', '+57').strip()
        telefono_raw = request.POST.get('telefono', '').strip()
        telefono = f'{codigo_pais}{telefono_raw}'
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '')
        password_confirm = request.POST.get('passwordConfirm', '')

        errors = []
        if not tipo_documento: errors.append('Selecciona un tipo de documento.')
        if not doc_usuario or not doc_usuario.isdigit() or not (5 <= len(doc_usuario) <= 15):
            errors.append('Número de documento inválido (5-15 dígitos).')
        if not primer_nombre: errors.append('El primer nombre es obligatorio.')
        if not primer_apellido: errors.append('El primer apellido es obligatorio.')
        if not telefono_raw or not telefono_raw.isdigit() or not (7 <= len(telefono_raw) <= 12):
            errors.append('Teléfono inválido (7-12 dígitos).')
        if not email or '@' not in email: errors.append('Email inválido.')
        if len(password) < 8: errors.append('La contraseña debe tener al menos 8 caracteres.')
        if not any(c.isupper() for c in password): errors.append('La contraseña debe contener al menos una mayúscula.')
        if not any(c.islower() for c in password): errors.append('La contraseña debe contener al menos una minúscula.')
        if not any(c.isdigit() for c in password): errors.append('La contraseña debe contener al menos un número.')
        if password != password_confirm: errors.append('Las contraseñas no coinciden.')

        if Usuario.objects.filter(email=email).exists():
            errors.append('Ya existe una cuenta con ese email.')
        if Usuario.objects.filter(doc_usuario=doc_usuario).exists():
            errors.append('Ya existe una cuenta con ese número de documento.')

        if errors:
            for error in errors:
                messages.error(request, error)
            return render(request, 'register.html', {'form': request.POST})

        # Map form values to model choices - PAS maps to PP (Pasaporte)
        tipo_doc_map = {
            'CC': TipoDocumento.CC, 
            'TI': TipoDocumento.TI, 
            'CE': TipoDocumento.CE, 
            'PAS': TipoDocumento.PP,
            'NIT': TipoDocumento.CC  # Default NIT to CC since no NIT option in model
        }
        tipo_doc = tipo_doc_map.get(tipo_documento, TipoDocumento.CC)

        rol_cliente, _ = Rol.objects.get_or_create(tipo_rol=TipoRol.CLIENTE, defaults={'nombre_rol': 'Cliente'})

        try:
            user = Usuario.objects.create_user(
                email=email, primer_nombre=primer_nombre, primer_apellido=primer_apellido,
                doc_usuario=doc_usuario, password=password, tipo_documento=tipo_doc,
                segundo_nombre=segundo_nombre, segundo_apellido=segundo_apellido,
                telefono=telefono, rol=rol_cliente,
            )
            messages.success(request, '¡Cuenta creada exitosamente! Por favor inicia sesión.')
            return redirect('/login/?registroExitoso=1')
        except Exception as e:
            messages.error(request, f'Error al crear la cuenta: {str(e)}')
            print(f'Registration error: {str(e)}')  # For debugging
            import traceback
            traceback.print_exc()  # For debugging
            return render(request, 'register.html', {'form': request.POST})

    return render(request, 'register.html')


@require_POST
def logout_view(request):
    """Cierra la sesión del usuario."""
    logout(request)
    return redirect('home')


# ── Vistas Administrativas (Dashboard) ──

@login_required
def usuario_list_view(request):
    """Vista para listar todos los usuarios con filtros y búsqueda + paginación flexible."""
    usuarios = _filter_usuarios_queryset(request)
    
    # Obtener items por página (valores permitidos: 10, 20, 30, 40, 50)
    items_por_pagina = int(request.GET.get('per_page', 10))
    if items_por_pagina not in [10, 20, 30, 40, 50]:
        items_por_pagina = 10
    
    # Crear paginador
    paginator = Paginator(usuarios, items_por_pagina)
    page_number = request.GET.get('page', 1)
    
    try:
        page_obj = paginator.page(page_number)
    except (PageNotAnInteger, EmptyPage):
        page_obj = paginator.page(1)
    
    # Construir query string para mantener filtros al navegar
    query_string = request.GET.copy()
    if 'page' in query_string:
        del query_string['page']
    query_params = query_string.urlencode()
    
    context = {
        'page_obj': page_obj,
        'roles': Rol.objects.all(),
        'total_usuarios': Usuario.objects.count(),
        'total_activos': Usuario.objects.filter(is_active=True).count(),
        'total_inactivos': Usuario.objects.filter(is_active=False).count(),
        'items_por_pagina': items_por_pagina,
        'query_params': query_params,
    }
    return render(request, 'dashboard/usuarios.html', context)


@admin_or_permission('add_usuario')
def usuario_create_view(request):
    """Vista para crear un nuevo usuario con hashing de contraseña."""
    if request.method == 'POST':
        form = UsuarioForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.set_password(form.cleaned_data['password'])
            user.save()
            messages.success(request, 'Usuario creado correctamente.')
            return redirect('usuarios_list')
    else:
        form = UsuarioForm()

    context = {
        'form': form,
        'title': 'Nuevo Usuario',
        'button_text': 'Crear Usuario',
        'roles': Rol.objects.all(),
    }
    return render(request, 'dashboard/form.html', context)


@admin_or_permission('change_usuario')
def usuario_update_view(request, pk):
    """Vista para editar un usuario con hashing opcional de contraseña."""
    user_instance = get_object_or_404(Usuario, pk=pk)
    
    if request.method == 'POST':
        form = UsuarioForm(request.POST, instance=user_instance)
        if form.is_valid():
            user = form.save(commit=False)
            new_password = form.cleaned_data.get('password')
            if new_password:
                user.set_password(new_password)
            user.save()
            messages.success(request, 'Usuario actualizado correctamente.')
            return redirect('usuarios_list')
    else:
        form = UsuarioForm(instance=user_instance)

    context = {
        'form': form,
        'usuario': user_instance,
        'title': f'Editar Usuario: {user_instance.email}',
        'button_text': 'Guardar Cambios',
        'roles': Rol.objects.all(),
    }
    return render(request, 'dashboard/form.html', context)


@admin_or_permission('delete_usuario')
def usuario_delete_view(request, pk):
    """Vista para eliminar un usuario."""
    usuario = get_object_or_404(Usuario, pk=pk)
    if request.method == 'POST':
        usuario.delete()
        messages.success(request, 'Usuario eliminado correctamente.')
        return redirect('usuarios_list')
    
    context = {
        'title': 'Eliminar Usuario',
        'object': usuario,
    }
    return render(request, 'dashboard/confirm_delete.html', context)


@login_required
def export_usuarios_view(request, file_format):
    """Vista para exportar usuarios en diferentes formatos (PDF, Excel, CSV)."""
    usuarios = _filter_usuarios_queryset(request)
    
    # ── EXPORTAR A PDF ──
    if file_format == 'pdf':
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
        from reportlab.lib import colors
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
        
        data = [['Nombre', 'Documento', 'Email', 'Teléfono', 'Rol', 'Estado']]
        for u in usuarios:
            data.append([
                u.nombre_completo,
                f"{u.get_tipo_documento_display()}: {u.doc_usuario}",
                u.email,
                u.telefono,
                u.rol.nombre_rol if u.rol else 'N/A',
                'Activo' if u.estado_usuario else 'Inactivo'
            ])
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E40AF')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
        ]))
        
        doc.build([table])
        buffer.seek(0)
        
        response = HttpResponse(buffer.read(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="usuarios.pdf"'
        return response
    
    # ── EXPORTAR A EXCEL ──
    elif file_format == 'xlsx':
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Usuarios"
        
        headers = ['Nombre Completo', 'Tipo Doc', 'Documento', 'Email', 'Teléfono', 'Rol', 'Estado', 'Fecha Creación']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for row, u in enumerate(usuarios, 2):
            ws.cell(row=row, column=1).value = u.nombre_completo
            ws.cell(row=row, column=2).value = u.get_tipo_documento_display()
            ws.cell(row=row, column=3).value = u.doc_usuario
            ws.cell(row=row, column=4).value = u.email
            ws.cell(row=row, column=5).value = u.telefono
            ws.cell(row=row, column=6).value = u.rol.nombre_rol if u.rol else 'N/A'
            ws.cell(row=row, column=7).value = 'Activo' if u.estado_usuario else 'Inactivo'
            ws.cell(row=row, column=8).value = u.created_at.strftime('%Y-%m-%d %H:%M')
        
        # Ajustar ancho
        for col in range(1, 9):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="usuarios.xlsx"'
        wb.save(response)
        return response
    
    # ── EXPORTAR A CSV ──
    elif file_format == 'csv':
        import csv
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="usuarios.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Nombre Completo', 'Tipo Documento', 'Documento', 'Email', 'Teléfono', 'Rol', 'Estado', 'Fecha Creación'])
        
        for u in usuarios:
            writer.writerow([
                u.nombre_completo,
                u.get_tipo_documento_display(),
                u.doc_usuario,
                u.email,
                u.telefono,
                u.rol.nombre_rol if u.rol else 'N/A',
                'Activo' if u.estado_usuario else 'Inactivo',
                u.created_at.strftime('%Y-%m-%d %H:%M')
            ])
        
        return response
    
    return redirect('usuarios_list')
