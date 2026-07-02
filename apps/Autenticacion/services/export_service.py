"""
Usuario Export Service - Handles export of usuario data in multiple formats.

This service provides:
- Multi-format export (PDF, XLSX, CSV)
- Data preparation and formatting
- File generation and HTTP response handling
"""

import csv
import logging
from io import BytesIO
from typing import Any, Dict, List, Optional

from django.http import HttpResponse
from django.db.models import QuerySet

from .base import ServiceBase
from ..models import Usuario

logger = logging.getLogger(__name__)


class UsuarioExportService(ServiceBase):
    """
    Service for exporting usuario data in various file formats.
    """
    
    # Export field configuration
    EXPORT_FIELDS = {
        'nombre_completo': 'Nombre Completo',
        'tipo_documento': 'Tipo Doc',
        'doc_usuario': 'Documento',
        'email': 'Email',
        'telefono': 'Teléfono',
        'rol': 'Rol',
        'estado': 'Estado',
        'created_at': 'Fecha Creación',
    }
    
    EXPORT_FIELDS_SIMPLE = {
        'nombre_completo': 'Nombre Completo',
        'tipo_documento': 'Tipo Documento',
        'doc_usuario': 'Documento',
        'email': 'Email',
        'rol': 'Rol',
        'estado': 'Estado',
    }
    
    @staticmethod
    def prepare_export_data(usuarios_queryset: QuerySet) -> List[Dict[str, Any]]:
        """
        Prepare usuario data for export in a consistent format.
        
        This method standardizes all the data that will be exported regardless
        of the target format (PDF, XLSX, CSV).
        
        Args:
            usuarios_queryset: QuerySet of Usuario objects to export
        
        Returns:
            List of dictionaries with formatted usuario data
        """
        data = []
        
        for usuario in usuarios_queryset:
            data.append({
                'nombre_completo': usuario.nombre_completo,
                'tipo_documento': usuario.get_tipo_documento_display() or 'N/A',
                'doc_usuario': usuario.doc_usuario,
                'email': usuario.email,
                'telefono': usuario.telefono or '',
                'rol': usuario.rol.nombre_rol if usuario.rol else 'N/A',
                'estado': 'Activo' if usuario.estado_usuario else 'Inactivo',
                'is_active': usuario.estado_usuario,  # For styling in some exports
                'created_at': usuario.created_at.strftime('%Y-%m-%d %H:%M') if usuario.created_at else '',
            })
        
        return data
    
    @staticmethod
    def export_to_pdf(usuarios_queryset: QuerySet, filename: str = 'usuarios.pdf') -> HttpResponse:
        """
        Export usuarios to PDF format using ReportLab.
        
        Creates a professional PDF report with:
        - Styled header row
        - Grid borders
        - Multiple columns of usuario data
        
        Args:
            usuarios_queryset: QuerySet of usuarios to export
            filename: Output filename
        
        Returns:
            HttpResponse with PDF file
        """
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
        except ImportError:
            logger.error("ReportLab not installed for PDF export")
            raise
        
        data = UsuarioExportService.prepare_export_data(usuarios_queryset)
        
        # Create PDF buffer
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=0.5*cm, bottomMargin=0.5*cm)
        
        # Build table data with headers
        table_data = [[
            'Nombre Completo',
            'Documento',
            'Email',
            'Teléfono',
            'Rol',
            'Estado'
        ]]
        
        for row in data:
            table_data.append([
                row['nombre_completo'],
                f"{row['tipo_documento']}: {row['doc_usuario']}",
                row['email'],
                row['telefono'],
                row['rol'],
                row['estado'],
            ])
        
        # Create table
        table = Table(table_data, repeatRows=1)
        
        # Apply styling
        table.setStyle(TableStyle([
            # Header row styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E40AF')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            # Data rows
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            # Alternating row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F4F6')]),
        ]))
        
        # Build PDF
        elements = [table]
        doc.build(elements)
        buffer.seek(0)
        
        UsuarioExportService.log_operation(
            'PDF export generated',
            {'filename': filename, 'record_count': len(data)}
        )
        
        response = HttpResponse(buffer.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    @staticmethod
    def export_to_xlsx(usuarios_queryset: QuerySet, filename: str = 'usuarios.xlsx') -> HttpResponse:
        """
        Export usuarios to Excel XML format using openpyxl.
        
        Creates an Excel spreadsheet with:
        - Styled header row (blue background, white text)
        - Frozen header row
        - Auto-adjusted column widths
        
        Args:
            usuarios_queryset: QuerySet of usuarios to export
            filename: Output filename
        
        Returns:
            HttpResponse with XLSX file
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            logger.error("openpyxl not installed for XLSX export")
            raise
        
        data = UsuarioExportService.prepare_export_data(usuarios_queryset)
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Usuarios"
        
        # Define headers
        headers = [
            'Nombre Completo',
            'Tipo Documento',
            'Documento',
            'Email',
            'Teléfono',
            'Rol',
            'Estado',
            'Fecha Creación'
        ]
        
        # Style for header row
        header_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Data alignment and styling
        data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB')
        )
        
        # Write data rows
        for row_idx, row_data in enumerate(data, 2):
            ws.cell(row=row_idx, column=1).value = row_data['nombre_completo']
            ws.cell(row=row_idx, column=2).value = row_data['tipo_documento']
            ws.cell(row=row_idx, column=3).value = row_data['doc_usuario']
            ws.cell(row=row_idx, column=4).value = row_data['email']
            ws.cell(row=row_idx, column=5).value = row_data['telefono']
            ws.cell(row=row_idx, column=6).value = row_data['rol']
            ws.cell(row=row_idx, column=7).value = row_data['estado']
            ws.cell(row=row_idx, column=8).value = row_data['created_at']
            
            # Apply formatting to data cells
            for col in range(1, 9):
                cell = ws.cell(row=row_idx, column=col)
                cell.alignment = data_alignment
                cell.border = thin_border
        
        # Apply header borders
        for col in range(1, 9):
            cell = ws.cell(row=1, column=col)
            cell.border = thin_border
        
        # Adjust column widths
        column_widths = {
            1: 25,  # Nombre Completo
            2: 16,  # Tipo Documento
            3: 16,  # Documento
            4: 25,  # Email
            5: 15,  # Teléfono
            6: 16,  # Rol
            7: 12,  # Estado
            8: 18,  # Fecha Creación
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Write to response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        
        UsuarioExportService.log_operation(
            'XLSX export generated',
            {'filename': filename, 'record_count': len(data)}
        )
        
        return response
    
    @staticmethod
    def export_to_csv(usuarios_queryset: QuerySet, filename: str = 'usuarios.csv') -> HttpResponse:
        """
        Export usuarios to CSV format.
        
        Creates a comma-separated values file with:
        - Header row with field names
        - All usuario data properly quoted
        
        Args:
            usuarios_queryset: QuerySet of usuarios to export
            filename: Output filename
        
        Returns:
            HttpResponse with CSV file
        """
        data = UsuarioExportService.prepare_export_data(usuarios_queryset)
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Add BOM for proper encoding in Excel
        response.write('\ufeff')
        
        writer = csv.writer(response, quoting=csv.QUOTE_ALL)
        
        # Write header
        writer.writerow([
            'Nombre Completo',
            'Tipo Documento',
            'Documento',
            'Email',
            'Teléfono',
            'Rol',
            'Estado',
            'Fecha Creación'
        ])
        
        # Write data rows
        for row in data:
            writer.writerow([
                row['nombre_completo'],
                row['tipo_documento'],
                row['doc_usuario'],
                row['email'],
                row['telefono'],
                row['rol'],
                row['estado'],
                row['created_at'],
            ])
        
        UsuarioExportService.log_operation(
            'CSV export generated',
            {'filename': filename, 'record_count': len(data)}
        )
        
        return response
    
    @staticmethod
    def export(
        usuarios_queryset: QuerySet,
        file_format: str = 'pdf',
        filename: str = None
    ) -> HttpResponse:
        """
        Unified export method that handles all formats.
        
        This is the main entry point for exporting usuario data.
        It dispatches to the appropriate format-specific method.
        
        Args:
            usuarios_queryset: QuerySet of usuarios to export
            file_format: Export format ('pdf', 'xlsx', 'csv')
            filename: Custom filename (optional)
        
        Returns:
            HttpResponse with file attachment
        
        Raises:
            ValueError: If file_format is not supported
        """
        file_format = file_format.lower().strip()
        
        if file_format == 'pdf':
            filename = filename or 'usuarios.pdf'
            return UsuarioExportService.export_to_pdf(usuarios_queryset, filename)
        
        elif file_format == 'xlsx':
            filename = filename or 'usuarios.xlsx'
            return UsuarioExportService.export_to_xlsx(usuarios_queryset, filename)
        
        elif file_format == 'csv':
            filename = filename or 'usuarios.csv'
            return UsuarioExportService.export_to_csv(usuarios_queryset, filename)
        
        else:
            error_msg = f'Unsupported export format: {file_format}'
            logger.error(error_msg)
            raise ValueError(error_msg)
